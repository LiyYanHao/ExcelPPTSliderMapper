import win32com.client
import os
import re
import json
import time
import psutil
import traceback
import pythoncom
from typing import Dict, List, Union, Any, Optional
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook


def check_file_in_use(file_path: str) -> bool:
    """检查文件是否被占用"""
    if not os.path.exists(file_path):
        return False
    
    try:
        # 尝试以独占方式打开文件
        with open(file_path, 'r+b') as f:
            pass
        return False
    except (OSError, IOError):
        return True

def close_powerpoint_processes():
    """关闭所有PowerPoint进程"""
    print("🔄 检查并关闭PowerPoint进程...")
    closed_count = 0
    
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.info['name'].lower() in ['powerpnt.exe', 'powerpoint.exe']:
                print(f"   发现PowerPoint进程 PID: {proc.info['pid']}")
                proc.terminate()
                proc.wait(timeout=5)
                closed_count += 1
                print(f"   已关闭PowerPoint进程 PID: {proc.info['pid']}")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.TimeoutExpired):
            pass
    
    if closed_count > 0:
        print(f"✅ 已关闭 {closed_count} 个PowerPoint进程")
        time.sleep(2)  # 等待进程完全关闭
    else:
        print("ℹ️  未发现需要关闭的PowerPoint进程")

def generate_unique_filename(base_path: str) -> str:
    """生成唯一的文件名"""
    if not os.path.exists(base_path):
        return base_path
    
    base_dir = os.path.dirname(base_path)
    base_name = os.path.splitext(os.path.basename(base_path))[0]
    extension = os.path.splitext(base_path)[1]
    
    counter = 1
    while True:
        new_path = os.path.join(base_dir, f"{base_name}_{counter}{extension}")
        if not os.path.exists(new_path):
            return new_path
        counter += 1

def safe_save_presentation(pres, output_path: str, max_retries: int = 3) -> bool:
    """安全保存演示文稿，带重试机制"""
    original_path = output_path
    
    for attempt in range(max_retries):
        try:
            print(f"💾 尝试保存文件 (第 {attempt + 1}/{max_retries} 次): {output_path}")
            
            # 检查文件是否被占用
            if check_file_in_use(output_path):
                print(f"⚠️  文件被占用: {output_path}")
                
                if attempt == 0:
                    # 第一次尝试：关闭PowerPoint进程
                    close_powerpoint_processes()
                    time.sleep(1)
                    continue
                else:
                    # 后续尝试：使用新文件名
                    output_path = generate_unique_filename(original_path)
                    print(f"🔄 使用新文件名: {output_path}")
            
            # 尝试保存
            pres.SaveAs(output_path)
            print(f"✅ 文件保存成功: {output_path}")
            return True
            
        except Exception as e:
            error_msg = str(e)
            print(f"❌ 保存失败 (第 {attempt + 1} 次): {error_msg}")
            
            if "正在使用中" in error_msg or "being used" in error_msg.lower():
                if attempt < max_retries - 1:
                    # 文件占用错误，尝试解决
                    print("🔄 检测到文件占用，尝试解决...")
                    close_powerpoint_processes()
                    
                    # 生成新文件名
                    output_path = generate_unique_filename(original_path)
                    print(f"🔄 使用新文件名: {output_path}")
                    time.sleep(2)
                    continue
            else:
                # 其他错误，直接重试
                if attempt < max_retries - 1:
                    print(f"⏳ 等待 {(attempt + 1) * 2} 秒后重试...")
                    time.sleep((attempt + 1) * 2)
                    continue
    
    print(f"💥 保存失败，已尝试 {max_retries} 次")
    return False


class PTMLParser:
    """PPT Template Markup Language (PTML) 解析器"""
    
    # 标记类型定义 - 使用 ${} 格式，只保留直接替换的标记
    MARKERS = {
        'NO_CONVERT': r'\$\{([A-Za-z][A-Za-z0-9_]*)\}',  # ${Value} - 不进行任何转换
    }

def print_slide_content(slide, page_num: int):
    """打印幻灯片的所有内容"""
    print(f"\n{'='*80}")
    print(f"第 {page_num} 页原始内容:")
    print(f"{'='*80}")
    
    if slide.Shapes.Count == 0:
        print("  该页面没有任何形状内容")
        return
    
    for shape_idx, shape in enumerate(slide.Shapes, 1):
        print(f"\n📍 形状 {shape_idx}:")
        print(f"   类型: {shape.Type} ({get_shape_type_name(shape.Type)})")
        print(f"   位置: Left={shape.Left:.1f}, Top={shape.Top:.1f}")
        print(f"   大小: Width={shape.Width:.1f}, Height={shape.Height:.1f}")
        
        # 处理文本框内容 - 改进版本
        if shape.HasTextFrame:
            text_frame = shape.TextFrame
            if text_frame.HasText:
                # 获取原始文本内容，不做任何处理
                text_content = text_frame.TextRange.Text
                
                # 显示完整的原始文本（不截断）
                print(f"   📝 文本内容 (长度: {len(text_content)}):")
                print(f"      原始文本: {repr(text_content)}")  # 使用repr显示所有字符
                print(f"      显示文本: 「{text_content}」")  # 正常显示
                
                # 如果文本很长，分行显示
                if len(text_content) > 100:
                    lines = text_content.split('\n')
                    print(f"      分行显示 ({len(lines)} 行):")
                    for i, line in enumerate(lines, 1):
                        if line.strip():  # 只显示非空行
                            print(f"        第{i}行: 「{line}」")
                
                # 详细的标记检测
                print(f"   🔍 标记检测结果:")
                all_markers = []
                
                # 逐个检测每种标记类型
                import re
                for marker_type, pattern in PTMLParser.MARKERS.items():
                    try:
                        matches = re.findall(pattern, text_content)
                        if matches:
                            print(f"      ✅ {marker_type}: {matches}")
                            all_markers.extend([(marker_type, match) for match in matches])
                        else:
                            print(f"      ❌ {marker_type}: 未找到")
                    except Exception as e:
                        print(f"      ⚠️  {marker_type}: 检测出错 - {e}")
                
                if not all_markers:
                    print(f"      ℹ️  未发现任何PTML标记")
                
                # 额外检查：查找所有可能的$标记
                simple_dollar_matches = re.findall(r'\$[^}]*\}?', text_content)
                if simple_dollar_matches:
                    print(f"   💡 发现的所有$标记: {simple_dollar_matches}")
                    
            else:
                print(f"   📝 文本框: 空内容")
        
        # 处理表格内容 - 改进版本
        if shape.HasTable:
            table = shape.Table
            print(f"   📊 表格: {table.Rows.Count} 行 × {table.Columns.Count} 列")
            
            # 打印完整的表格内容，不截断
            print(f"      完整表格内容:")
            for row in range(1, table.Rows.Count + 1):
                row_data = []
                for col in range(1, table.Columns.Count + 1):
                    try:
                        cell = table.Cell(row, col)
                        if cell.Shape.HasTextFrame and cell.Shape.TextFrame.HasText:
                            cell_text = cell.Shape.TextFrame.TextRange.Text
                            # 不截断，显示完整内容，并检查标记
                            row_data.append(f"「{cell_text}」")
                            
                            # 检查单元格中的标记
                            cell_markers = []
                            for marker_type, pattern in PTMLParser.MARKERS.items():
                                matches = re.findall(pattern, cell_text)
                                if matches:
                                    cell_markers.extend([(marker_type, match) for match in matches])
                            
                            if cell_markers:
                                print(f"        单元格[{row},{col}]原文: {repr(cell_text)}")
                                print(f"        单元格[{row},{col}]标记: {cell_markers}")
                                
                        else:
                            row_data.append("「空」")
                    except Exception as e:
                        row_data.append(f"「错误:{e}」")
                
                if row == 1:
                    print(f"        表头: {' | '.join(row_data)}")
                else:
                    print(f"        第{row}行: {' | '.join(row_data)}")
        
        # 处理图表内容 - 改进版本
        if shape.HasChart:
            chart = shape.Chart
            chart_type_name = get_chart_type_name(chart.ChartType)
            print(f"   📈 图表: {chart_type_name} (类型码: {chart.ChartType})")
            
            # 获取图表标题
            try:
                if chart.HasTitle:
                    title_text = chart.ChartTitle.Text
                    print(f"      标题原文: {repr(title_text)}")
                    print(f"      标题显示: 「{title_text}」")
                    
                    # 检查标题中的标记
                    title_markers = []
                    for marker_type, pattern in PTMLParser.MARKERS.items():
                        matches = re.findall(pattern, title_text)
                        if matches:
                            title_markers.extend([(marker_type, match) for match in matches])
                    
                    if title_markers:
                        print(f"      标题标记: {title_markers}")
                else:
                    print(f"      标题: 无")
            except Exception as e:
                print(f"      标题: 读取出错 - {e}")
            
            # 获取数据系列信息
            try:
                series_count = chart.SeriesCollection.Count
                print(f"      数据系列数量: {series_count}")
                
                for i in range(1, series_count + 1):  # 显示所有系列
                    try:
                        series = chart.SeriesCollection(i)
                        series_name = str(series.Name)
                        print(f"      系列 {i} 原文: {repr(series_name)}")
                        print(f"      系列 {i} 显示: 「{series_name}」")
                        
                        # 检查系列名中的标记
                        series_markers = []
                        for marker_type, pattern in PTMLParser.MARKERS.items():
                            matches = re.findall(pattern, series_name)
                            if matches:
                                series_markers.extend([(marker_type, match) for match in matches])
                        
                        if series_markers:
                            print(f"      系列 {i} 标记: {series_markers}")
                            
                    except Exception as e:
                        print(f"      系列 {i}: 读取出错 - {e}")
                        
            except Exception as e:
                print(f"      数据系列: 读取出错 - {e}")
        
        # 处理图片内容 - 改进版本
        if shape.Type == 13:  # 图片类型
            print(f"   🖼️  图片:")
            try:
                if hasattr(shape, 'Name'):
                    shape_name = shape.Name
                    print(f"      名称: 「{shape_name}」")
                    
                if hasattr(shape, 'AlternativeText'):
                    alt_text = shape.AlternativeText
                    if alt_text:
                        print(f"      替代文本原文: {repr(alt_text)}")
                        print(f"      替代文本显示: 「{alt_text}」")
                        
                        # 检查替代文本中的标记
                        alt_markers = []
                        for marker_type, pattern in PTMLParser.MARKERS.items():
                            matches = re.findall(pattern, alt_text)
                            if matches:
                                alt_markers.extend([(marker_type, match) for match in matches])
                        
                        if alt_markers:
                            print(f"      替代文本标记: {alt_markers}")
                    else:
                        print(f"      替代文本: 无")
                        
            except Exception as e:
                print(f"      图片信息: 读取出错 - {e}")
        
        # 处理其他形状 - 改进版本
        if not shape.HasTextFrame and not shape.HasTable and not shape.HasChart and shape.Type != 13:
            try:
                shape_name = getattr(shape, 'Name', '未知')
                print(f"   🔹 其他形状: 「{shape_name}」")
                
                # 尝试获取其他可能的文本属性
                if hasattr(shape, 'TextFrame2'):
                    try:
                        if shape.TextFrame2.HasText:
                            text_content = shape.TextFrame2.TextRange.Text
                            print(f"      TextFrame2内容原文: {repr(text_content)}")
                            print(f"      TextFrame2内容显示: 「{text_content}」")
                            
                            # 检查TextFrame2中的标记
                            tf2_markers = []
                            for marker_type, pattern in PTMLParser.MARKERS.items():
                                matches = re.findall(pattern, text_content)
                                if matches:
                                    tf2_markers.extend([(marker_type, match) for match in matches])
                            
                            if tf2_markers:
                                print(f"      TextFrame2标记: {tf2_markers}")
                    except:
                        pass
                        
            except Exception as e:
                print(f"   🔹 其他形状: 信息读取出错 - {e}")
    
    print(f"\n{'='*80}")
    print(f"第 {page_num} 页内容扫描完成 - 共 {slide.Shapes.Count} 个形状")
    print(f"{'='*80}")

def get_shape_type_name(shape_type: int) -> str:
    """获取形状类型的中文名称"""
    shape_types = {
        1: "自选图形",
        2: "标注",
        3: "图表",
        4: "注释",
        5: "自由曲线",
        6: "组合",
        7: "嵌入式OLE对象",
        8: "窗体控件",
        9: "线条",
        10: "链接式OLE对象",
        11: "链接式图片",
        12: "媒体",
        13: "图片",
        14: "占位符",
        15: "多边形",
        16: "多段线",
        17: "文本框",
        18: "表格",
        19: "文本效果"
    }
    return shape_types.get(shape_type, f"未知类型({shape_type})")

def get_chart_type_name(chart_type: int) -> str:
    """获取图表类型的中文名称"""
    chart_types = {
        4: "折线图",
        5: "饼图",
        51: "柱状图",
        52: "堆积柱状图",
        53: "百分比堆积柱状图",
        57: "条形图",
        65: "面积图",
        68: "散点图",
        69: "气泡图",
        70: "圆环图",
        72: "雷达图"
    }
    return chart_types.get(chart_type, f"未知图表({chart_type})")

def process_ptml_template(ppt_path: str, template_data: Dict[str, Any], output_path: Optional[str] = None, page_numbers: Optional[List[int]] = None) -> bool:
    """处理PPT模板，替换其中的标记"""
    import win32com.client
    import os
    import time
    
    # 确保输入文件存在
    if not os.path.exists(ppt_path):
        print(f"❌ 输入文件不存在: {ppt_path}")
        return False
    
    # 如果没有指定输出路径，使用输入路径
    if output_path is None:
        output_path = ppt_path
    
    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
            print(f"✅ 创建输出目录: {output_dir}")
        except Exception as e:
            print(f"❌ 创建输出目录失败: {e}")
            return False
    
    print(f"🔄 开始处理PPT模板: {ppt_path}")
    
    try:
        # 创建PowerPoint应用程序实例
        ppt = win32com.client.Dispatch("PowerPoint.Application")
        ppt.Visible = True  # 设置为可见，方便调试
        
        try:
            # 打开源文件
            print(f"📂 正在打开文件: {ppt_path}")
            source_pres = ppt.Presentations.Open(ppt_path)
            print(f"✅ 文件打开成功")
            
            # 处理所有幻灯片
            total_slides = source_pres.Slides.Count
            print(f"📊 总页数: {total_slides}")
            
            # 如果指定了页码，只处理指定页码
            slides_to_process = page_numbers if page_numbers else range(1, total_slides + 1)
            
            for page_num in slides_to_process:
                if page_num > total_slides:
                    print(f"⚠️ 跳过无效页码: {page_num} (超出总页数)")
                    continue
                
                print(f"\n处理第 {page_num} 页:")
                slide = source_pres.Slides(page_num)
                
                # 打印幻灯片内容（调试用）
                print_slide_content(slide, page_num)
                
                # 处理幻灯片中的所有形状
                for shape_idx, shape in enumerate(slide.Shapes, 1):
                    print(f"\n处理形状 {shape_idx}:")
                    
                    # 根据形状类型调用相应的处理函数
                    if shape.Type == 6:  # 组合形状
                        process_group_shape(shape, template_data)
                    elif shape.HasTable:
                        process_table_shape(shape, template_data)
                    elif shape.HasChart:
                        process_chart_shape(shape, template_data)
                    elif shape.Type == 13:  # 图片
                        process_image_shape(shape, template_data)
                    else:
                        process_text_shape(shape, template_data)
            
            # 保存文件
            print(f"\n💾 正在保存文件: {output_path}")
            if safe_save_presentation(source_pres, output_path):
                print(f"✅ 文件保存成功: {output_path}")
                return True
            else:
                print(f"❌ 文件保存失败")
                return False
                
        finally:
            try:
                # 关闭演示文稿
                source_pres.Close()
                print("✅ 已关闭演示文稿")
            except:
                pass
            
            try:
                # 退出PowerPoint
                ppt.Quit()
                print("✅ 已退出PowerPoint")
            except:
                pass
            
            # 确保所有PowerPoint进程都被关闭
            print("🔄 检查并关闭PowerPoint进程...")
            close_powerpoint_processes()
            
    except Exception as e:
        print(f"创建PowerPoint应用程序实例时出错: {e}")
        print(f"详细错误信息: {traceback.format_exc()}")
        return False
    
    return True

def process_group_shape(shape, template_data: Dict[str, Any]):
    """处理组合形状中的标记（递归处理每个子形状）"""
    if shape.Type != 6:  # 不是组合形状
        return
    
    print(f"  📦 处理组合形状: {shape.Name}")
    
    try:
        # 先检查组合形状本身是否有文本
        if shape.HasTextFrame:
            print(f"    📝 组合形状本身有文本框")
            process_text_shape(shape, template_data)
        
        # 递归处理组合中的每个子形状
        for sub_shape_idx, sub_shape in enumerate(shape.GroupItems, 1):
            print(f"    🔧 处理子形状 {sub_shape_idx}: {get_shape_type_name(sub_shape.Type)}")
            
            # 递归处理嵌套的组合形状
            if sub_shape.Type == 6:  # 嵌套的组合形状
                process_group_shape(sub_shape, template_data)
            
            # 处理子形状的文本
            if sub_shape.HasTextFrame:
                print(f"      📝 子形状有文本框，开始处理...")
                process_text_shape(sub_shape, template_data)
            
            # 处理子形状的表格
            if sub_shape.HasTable:
                process_table_shape(sub_shape, template_data)
            
            # 处理子形状的图表
            if sub_shape.HasChart:
                process_chart_shape(sub_shape, template_data)
            
            # 处理子形状的图片
            if sub_shape.Type == 13:  # 图片形状
                process_image_shape(sub_shape, template_data)
                
    except Exception as e:
        print(f"    ⚠️  处理组合形状时出错: {e}")
        import traceback
        print(f"    详细错误: {traceback.format_exc()}")

def get_case_insensitive_value(key: str, data_dict: Dict[str, Any], key_mapping: Dict[str, str]) -> Any:
    """获取不区分大小写的键值"""
    # 先尝试直接匹配
    if key in data_dict:
        return data_dict[key]
    
    # 如果直接匹配失败，尝试不区分大小写匹配
    if key.upper() in key_mapping:
        original_key = key_mapping[key.upper()]
        if original_key in data_dict:
            return data_dict[original_key]
    
    # 如果都没有匹配到，返回None
    return None


def process_text_shape(shape, template_data: Dict[str, Any]):
    """处理文本形状，直接替换标记值，不进行任何转换"""
    if not shape.HasTextFrame:
        return
    
    text_frame = shape.TextFrame
    if not text_frame.HasText:
        return
    
    original_text = text_frame.TextRange.Text
    modified_text = original_text
    print(f"  🔍 处理文本: '{original_text}'")
    
    # 获取key映射字典
    key_mapping = template_data.get("_key_mapping", {})
    
    # 处理所有标记，直接替换不进行转换
    no_convert_markers = re.findall(PTMLParser.MARKERS['NO_CONVERT'], modified_text)
    if no_convert_markers:
        print(f"  发现标记: {no_convert_markers}")
        for marker in no_convert_markers:
            value = get_case_insensitive_value(marker, template_data.get('TEXT', {}), key_mapping.get("TEXT", {}))
            if value is not None:
                modified_text = modified_text.replace(f"${{{marker}}}", str(value))
                print(f"    直接替换: ${{{marker}}} -> {value}")
            else:
                print(f"    未找到匹配: ${{{marker}}}")
    
    # 如果文本有变化，则更新
    if modified_text != original_text:
        text_frame.TextRange.Text = modified_text
        print(f"  ✅ 文本已更新")

def process_table_shape(shape, template_data: Dict[str, Any]):
    """处理表格形状中的标记，直接替换不进行转换"""
    if not shape.HasTable:
        return
    
    try:
        table = shape.Table
        print(f"  处理表格: {table.Rows.Count} 行 x {table.Columns.Count} 列")
        
        # 获取key映射字典
        key_mapping = template_data.get("_key_mapping", {})
        
        # 遍历表格的每个单元格
        for row in range(1, table.Rows.Count + 1):
            for col in range(1, table.Columns.Count + 1):
                try:
                    cell = table.Cell(row, col)
                    if cell.Shape.HasTextFrame and cell.Shape.TextFrame.HasText:
                        cell_text = cell.Shape.TextFrame.TextRange.Text
                        original_text = cell_text
                        
                        # 处理所有标记，直接替换不进行转换
                        no_convert_markers = re.findall(PTMLParser.MARKERS['NO_CONVERT'], cell_text)
                        if no_convert_markers:
                            print(f"    发现标记: {no_convert_markers}")
                            for marker in no_convert_markers:
                                value = get_case_insensitive_value(marker, template_data.get('TEXT', {}), key_mapping.get("TEXT", {}))
                                if value is not None:
                                    cell_text = cell_text.replace(f"${{{marker}}}", str(value))
                                    print(f"      直接替换: ${{{marker}}} -> {value}")
                                else:
                                    print(f"      未找到匹配: ${{{marker}}}")
                            
                            # 只有在文本有变化时才更新
                            if cell_text != original_text:
                                cell.Shape.TextFrame.TextRange.Text = cell_text
                                print(f"      ✅ 单元格[{row},{col}]已更新")
                except Exception as e:
                    print(f"    处理单元格[{row},{col}]时出错: {e}")
    except Exception as e:
        print(f"    处理表格时出错: {e}")
        print(f"    详细错误信息: {traceback.format_exc()}")

def process_chart_shape(shape, template_data: Dict[str, Any]):
    """处理图表形状中的标记，包括组合图表"""
    if not shape.HasChart:
        return
    
    chart = shape.Chart
    print(f"  处理图表: {chart.ChartType}")
    
    # 获取key映射字典
    key_mapping = template_data.get("_key_mapping", {})
    
    try:
        # 处理图表标题
        try:
            if chart.HasTitle:
                title_text = chart.ChartTitle.Text
                original_text = title_text
                
                # 处理所有标记，直接替换不进行转换
                no_convert_markers = re.findall(PTMLParser.MARKERS['NO_CONVERT'], title_text)
                if no_convert_markers:
                    print(f"    发现标记: {no_convert_markers}")
                    for marker in no_convert_markers:
                        value = get_case_insensitive_value(marker, template_data.get('TEXT', {}), key_mapping.get("TEXT", {}))
                        if value is not None:
                            title_text = title_text.replace(f"${{{marker}}}", str(value))
                            print(f"      直接替换: ${{{marker}}} -> {value}")
                        else:
                            print(f"      未找到匹配: ${{{marker}}}")
                    
                    # 只有在文本有变化时才更新
                    if title_text != original_text:
                        chart.ChartTitle.Text = title_text
                        print(f"      ✅ 图表标题已更新")
        except Exception as e:
            print(f"    处理图表标题时出错: {e}")
        
        # 处理图表数据更新
        try:
            # 检查是否有对应的图表数据
            chart_markers = re.findall(PTMLParser.MARKERS['NO_CONVERT'], str(shape.AlternativeText) if hasattr(shape, 'AlternativeText') else "")
            
            for marker in chart_markers:
                chart_data = get_case_insensitive_value(marker, template_data.get('CHARTS', {}), key_mapping.get("CHARTS", {}))
                if chart_data and isinstance(chart_data, dict):
                    print(f"    更新图表数据: {marker}")
                    update_chart_data(chart, chart_data)
                    
        except Exception as e:
            print(f"    处理图表数据时出错: {e}")
            
    except Exception as e:
        print(f"  处理图表时出错: {e}")

def update_chart_data(chart, chart_data: Dict[str, Any]):
    """更新图表数据，支持组合图表"""
    try:
        chart_type = chart_data.get("type", "column")
        categories = chart_data.get("categories", [])
        
        print(f"    📊 更新图表类型: {chart_type}")
        print(f"    📊 分类数量: {len(categories)}")
        
        # 更新图表标题
        if chart_data.get("title") and chart.HasTitle:
            chart.ChartTitle.Text = chart_data["title"]
            print(f"    ✅ 图表标题已更新为: {chart_data['title']}")
        
        # 处理组合图表
        if chart_type.lower() == "combo":
            column_series = chart_data.get("column_series", [])
            line_series = chart_data.get("line_series", [])
            
            print(f"    📊 柱状图系列: {len(column_series)} 个")
            print(f"    📊 折线图系列: {len(line_series)} 个")
            
            # 更新数据表（工作表）
            try:
                chart_workbook = chart.ChartData.Workbook
                chart_worksheet = chart_workbook.Worksheets(1)
                
                # 清空现有数据
                chart_worksheet.UsedRange.Clear()
                
                # 设置分类（X轴标签）
                for i, category in enumerate(categories, 2):  # 从第2行开始
                    chart_worksheet.Cells(i, 1).Value = category
                
                # 设置柱状图数据
                col_idx = 2  # 从第2列开始
                for series_name in set(item["name"] for item in column_series):
                    chart_worksheet.Cells(1, col_idx).Value = series_name
                    
                    # 按分类组织数据
                    for i, category in enumerate(categories, 2):
                        # 查找该分类对应的值
                        value = 0
                        for item in column_series:
                            if item["category"] == category and item["name"] == series_name:
                                value = item["value"]
                                break
                        chart_worksheet.Cells(i, col_idx).Value = value
                    
                    col_idx += 1
                
                # 设置折线图数据
                for series_name in set(item["name"] for item in line_series):
                    chart_worksheet.Cells(1, col_idx).Value = series_name
                    
                    # 按分类组织数据
                    for i, category in enumerate(categories, 2):
                        # 查找该分类对应的值
                        value = 0
                        for item in line_series:
                            if item["category"] == category and item["name"] == series_name:
                                value = item["value"]
                                break
                        chart_worksheet.Cells(i, col_idx).Value = value
                    
                    col_idx += 1
                
                print(f"    ✅ 图表数据已更新")
                
                # 设置图表系列类型
                try:
                    series_count = chart.SeriesCollection().Count
                    column_series_count = len(set(item["name"] for item in column_series))
                    
                    # 设置前面的系列为柱状图
                    for i in range(1, min(column_series_count + 1, series_count + 1)):
                        series = chart.SeriesCollection(i)
                        series.ChartType = 51  # xlColumnClustered
                    
                    # 设置后面的系列为折线图，并使用次坐标轴
                    for i in range(column_series_count + 1, series_count + 1):
                        series = chart.SeriesCollection(i)
                        series.ChartType = 4   # xlLine
                        series.AxisGroup = 2   # 次坐标轴
                    
                    print(f"    ✅ 图表系列类型已设置")
                    
                except Exception as e:
                    print(f"    ⚠️  设置图表系列类型时出错: {e}")
                
            except Exception as e:
                print(f"    ⚠️  更新图表数据时出错: {e}")
        
        else:
            # 处理其他类型的图表
            print(f"    ℹ️  暂不支持图表类型: {chart_type}")
            
    except Exception as e:
        print(f"    ❌ 更新图表数据失败: {e}")
        import traceback
        print(f"    详细错误: {traceback.format_exc()}")

def process_image_shape(shape, template_data: Dict[str, Any]):
    """处理图片形状中的标记"""
    if shape.Type != 13:  # 不是图片形状
        return
    
    # 获取key映射字典
    key_mapping = template_data.get("_key_mapping", {})
    
    # 检查图片的替代文本中是否有标记
    try:
        if hasattr(shape, 'AlternativeText'):
            image_text = shape.AlternativeText
            image_markers = re.findall(PTMLParser.MARKERS['IMAGE'], image_text)
            if image_markers:
                print(f"  发现图片标记: {image_markers}")
                for marker in image_markers:
                    new_image_path = get_case_insensitive_value(marker, template_data.get('IMAGES', {}), key_mapping.get("IMAGES", {}))
                    if new_image_path is not None and os.path.exists(new_image_path):
                        # 记录原图片位置信息
                        left, top = shape.Left, shape.Top
                        width, height = shape.Width, shape.Height
                        slide = shape.Parent
                        
                        # 删除原图片
                        shape.Delete()
                        
                        # 添加新图片
                        slide.Shapes.AddPicture(
                            new_image_path,
                            False, True,
                            left, top, width, height
                        )
                        print(f"    已替换图片: {new_image_path}")
    except Exception as e:
        print(f"    处理图片标记时出错: {e}")

def read_excel_template(excel_path: str) -> Dict[str, Any]:
    """从Excel模板中读取数据，保持原始格式（包括百分比）"""
    try:
        print(f"\n📊 读取Excel模板: {excel_path}")
        
        # 使用openpyxl直接读取Excel文件
        wb = load_workbook(excel_path)
        
        template_data = {
            "TEXT": {},
            "DATES": {},
            "TABLES": {},
            "CHARTS": {},
            "IMAGES": {},
            "CONDITIONS": {}
        }
        
        # 创建key映射字典，用于存储大小写映射关系
        key_mapping = {
            "TEXT": {},
            "DATES": {},
            "TABLES": {},
            "CHARTS": {},
            "IMAGES": {},
            "CONDITIONS": {}
        }
        
        # 处理每个sheet
        for sheet_name in wb.sheetnames:
            print(f"\n📑 处理工作表: {sheet_name}")
            
            if sheet_name.lower() == "text":
                sheet = wb[sheet_name]
                
                # 找到key和value的列索引
                header_row = next(sheet.rows)
                key_col = None
                value_col = None
                for idx, cell in enumerate(header_row, 1):
                    if cell.value and str(cell.value).lower() == 'key':
                        key_col = idx
                    elif cell.value and str(cell.value).lower() == 'value':
                        value_col = idx
                
                if key_col is None or value_col is None:
                    print(f"  ⚠️ 在工作表 {sheet_name} 中未找到必要的列")
                    continue
                
                # 从第二行开始处理数据（跳过标题行）
                for row in list(sheet.rows)[1:]:
                    key_cell = row[key_col - 1]
                    value_cell = row[value_col - 1]
                    
                    if not key_cell.value:
                        continue
                    
                    key = str(key_cell.value).strip()
                    
                    # 根据单元格格式处理值
                    if value_cell.number_format and '%' in value_cell.number_format:
                        # 如果是百分比格式，直接使用原始字符串值
                        try:
                            # 获取单元格的原始值
                            raw_value = value_cell.value
                            # 如果是数字类型，转换为整数百分比格式
                            if isinstance(raw_value, (int, float)):
                                # 将数值乘以100并四舍五入为整数
                                percentage = round(raw_value * 100)
                                value = f"{percentage}%"
                            else:
                                # 如果不是数字，尝试从字符串中提取数字
                                value = str(raw_value or '').strip()
                                if value:
                                    # 如果字符串中包含数字，尝试转换为整数百分比
                                    try:
                                        # 移除所有非数字字符（保留负号）
                                        num_str = ''.join(c for c in value if c.isdigit() or c == '-')
                                        if num_str:
                                            num_value = round(float(num_str))
                                            value = f"{num_value}%"
                                        elif not value.endswith('%'):
                                            value = f"{value}%"
                                    except ValueError:
                                        # 如果转换失败，保持原始值
                                        if not value.endswith('%'):
                                            value = f"{value}%"
                        except Exception as e:
                            print(f"  ⚠️  处理百分比值时出错: {e}")
                            value = str(value_cell.value or '').strip()
                            if value and not value.endswith('%'):
                                value = f"{value}%"
                    else:
                        # 其他情况使用显示值
                        value = str(value_cell.value or '').strip()
                    
                    # 保存原始key和大写key的映射关系
                    key_mapping["TEXT"][key.upper()] = key
                    template_data["TEXT"][key] = value
                    print(f"  📝 文本: {key} -> {value}")
            
            elif sheet_name.lower() == "dates":
                # 处理日期数据
                sheet = wb[sheet_name]
                for row in list(sheet.rows)[1:]:  # 跳过标题行
                    if not row[0].value:  # 检查key是否存在
                        continue
                    
                    key = str(row[0].value).strip()
                    value = row[1].value
                    
                    if isinstance(value, (datetime, str)):
                        if isinstance(value, datetime):
                            value = value.strftime('%Y-%m-%d')
                        else:
                            value = value.strip()
                        
                        key_mapping["DATES"][key.upper()] = key
                        template_data["DATES"][key] = value
                        print(f"  📅 日期: {key} -> {value}")
            
            elif sheet_name.lower() == "combo_charts":
                # 处理组合图表数据（柱状图+折线图）
                sheet = wb[sheet_name]
                current_chart = None
                chart_data = {}
                
                for row in list(sheet.rows)[1:]:  # 跳过标题行
                    if not row[0].value:
                        continue
                    
                    chart_name = str(row[0].value).strip()
                    category = str(row[1].value or '').strip()
                    series_type = str(row[2].value or '').strip()
                    series_name = str(row[3].value or '').strip()
                    value = row[4].value
                    chart_type = str(row[5].value or '').strip()
                    title = str(row[6].value or '').strip()
                    
                    # 初始化图表数据结构
                    if chart_name not in chart_data:
                        chart_data[chart_name] = {
                            "type": chart_type or "combo",
                            "title": title,
                            "categories": [],
                            "column_series": [],  # 柱状图系列
                            "line_series": []     # 折线图系列
                        }
                    
                    # 添加分类
                    if category and category not in chart_data[chart_name]["categories"]:
                        chart_data[chart_name]["categories"].append(category)
                    
                    # 添加数据系列
                    if series_type and series_name and value is not None:
                        series_data = {
                            "name": series_name,
                            "category": category,
                            "value": float(value) if isinstance(value, (int, float)) else 0
                        }
                        
                        if series_type.lower() == "column":
                            chart_data[chart_name]["column_series"].append(series_data)
                        elif series_type.lower() == "line":
                            chart_data[chart_name]["line_series"].append(series_data)
                
                # 保存到模板数据中
                for chart_name, data in chart_data.items():
                    key_mapping["CHARTS"][chart_name.upper()] = chart_name
                    template_data["CHARTS"][chart_name] = data
                    print(f"  📊 组合图表: {chart_name}")
                    print(f"    类型: {data['type']}")
                    print(f"    标题: {data['title']}")
                    print(f"    分类: {data['categories']}")
                    print(f"    柱状图系列: {len(data['column_series'])} 个")
                    print(f"    折线图系列: {len(data['line_series'])} 个")
            
            elif sheet_name.lower() == "revenue_data":
                # 处理收入数据表格
                sheet = wb[sheet_name]
                revenue_table_data = {
                    "headers": [],
                    "data": []
                }
                
                # 读取表头
                header_row = next(sheet.rows)
                for cell in header_row:
                    if cell.value:
                        revenue_table_data["headers"].append(str(cell.value))
                
                # 读取数据行
                for row in list(sheet.rows)[1:]:  # 跳过标题行
                    row_data = []
                    for cell in row:
                        if cell.value is not None:
                            row_data.append(str(cell.value))
                        else:
                            row_data.append("")
                    if any(row_data):  # 如果行中有数据
                        revenue_table_data["data"].append(row_data)
                
                # 保存到模板数据中
                key_mapping["TABLES"]["REVENUE_DATA"] = "revenue_data"
                template_data["TABLES"]["revenue_data"] = revenue_table_data
                print(f"  📊 收入数据表格: {len(revenue_table_data['data'])} 行数据")
                print(f"    表头: {revenue_table_data['headers']}")
            
            elif sheet_name.lower() == "tables":
                # 处理表格数据
                sheet = wb[sheet_name]
                current_table = None
                headers = []
                data = []
                
                for row in sheet.rows:
                    if not row[0].value:
                        continue
                    
                    first_cell = str(row[0].value).strip()
                    if first_cell.lower() == 'table_name':
                        if current_table and headers:
                            template_data["TABLES"][current_table] = {
                                "headers": headers,
                                "data": data
                            }
                        current_table = str(row[1].value).strip()
                        headers = []
                        data = []
                    elif first_cell.lower() == 'header':
                        headers = [str(cell.value).strip() for cell in row[1:] if cell.value]
                    else:
                        row_data = []
                        for cell in row:
                            if cell.value is None:
                                row_data.append("")
                            else:
                                row_data.append(str(cell.value).strip())
                        if any(row_data):
                            data.append(row_data)
                
                if current_table and headers:
                    template_data["TABLES"][current_table] = {
                        "headers": headers,
                        "data": data
                    }
            
            elif sheet_name.lower() == "images":
                # 处理图片路径
                sheet = wb[sheet_name]
                for row in list(sheet.rows)[1:]:  # 跳过标题行
                    if not row[0].value or not row[1].value:
                        continue
                    
                    key = str(row[0].value).strip()
                    path = str(row[1].value).strip()
                    
                    key_mapping["IMAGES"][key.upper()] = key
                    template_data["IMAGES"][key] = path
                    print(f"  🖼️  图片: {key} -> {path}")
        
        # 将key映射添加到template_data中
        template_data["_key_mapping"] = key_mapping
        print("\n✅ Excel模板数据读取完成")
        return template_data
        
    except Exception as e:
        print(f"\n❌ 读取Excel模板时出错: {e}")
        import traceback
        print(f"详细错误信息: {traceback.format_exc()}")
        raise


if __name__ == "__main__":
    # 文件路径
    ppt_file = r"D:\pythonProject\LanchainProject\tests\ppt_chuli\无仓年度PPT-模版.pptx"
    excel_template = r"D:\pythonProject\LanchainProject\tests\ppt_chuli\template_data.xlsx"  # Excel模板路径
    output_path = r"D:\pythonProject\LanchainProject\tests\ppt_chuli\生成的单页报告.pptx"
    
    # 从Excel读取模板数据
    template_data = read_excel_template(excel_template)
    
    # 处理PPT
    process_ptml_template(ppt_file, template_data, output_path, page_numbers=[1,2,3,4,5,6])